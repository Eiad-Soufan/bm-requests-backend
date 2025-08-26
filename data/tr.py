#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import csv
import os
import re
import sys
import unicodedata
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ تحتاج openpyxl:  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# --------- إعدادات أعمدة مرنة ----------
CODE_HEADERS = [
    "serial number", "code", "رمز", "الكود", "رقم النموذج",
    "form code", "model code"
]
AR_HEADERS = ["name (arabic)", "arabic name", "name_ar", "الاسم بالعربية", "الاسم العربي"]
EN_HEADERS = ["name (english)", "english name", "name_en", "الاسم بالانجليزي", "الاسم الإنجليزي"]
FILENAME_HEADERS = ["file_name", "filename", "file", "pdf", "pdf_name", "اسم الملف"]

PDF_EXT = ".pdf"

# خريطة تحويل الأرقام العربية -> الإنجليزية
AR_DIGITS = "٠١٢٣٤٥٦٧٨٩"
EN_DIGITS = "0123456789"
TRANS_DIGITS = str.maketrans({a: b for a, b in zip(AR_DIGITS, EN_DIGITS)})

# إزالة التشكيل والمدّ
AR_TATWEEL = "\u0640"


def normalize_key(s: str) -> str:
    """
    تطبيع نص للمقارنة:
    - NFKC
    - تحويل للأحرف الصغيرة
    - تحويل أرقام عربية لإنجليزية
    - إزالة التشكيل والمدّ
    - إزالة الامتداد .pdf إن وُجد
    - إزالة كل ما ليس حرفًا/رقمًا
    """
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\.pdf$", "", s, flags=re.IGNORECASE)
    s = unicodedata.normalize("NFKC", s).lower()
    s = s.translate(TRANS_DIGITS)
    # إزالة المدّ
    s = s.replace(AR_TATWEEL, "")
    # إزالة التشكيل (Combining marks)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    # أبقِ فقط حروف/أرقام (عربي + لاتيني)
    s = re.sub(r"[^0-9a-z\u0600-\u06FF]+", "", s)
    return s


def find_col_idx(header_cells, candidates):
    """يعيد رقم العمود (1-based) لأول اسم متاح من candidates، أو None"""
    header = [normalize_key(c.value) if c.value else "" for c in header_cells]
    cands_norm = [normalize_key(x) for x in candidates]
    for i, cell in enumerate(header, start=1):
        if cell in cands_norm:
            return i
    return None


def build_mapping_from_excel(excel_path: Path):
    """
    يبني خرائط:
      - name_key (عربي/إنجليزي/اسم ملف) -> code (الأصلي)
      - code_key -> code (للتعرف على الملفات التي اسمها فعلاً كود)
    """
    wb = load_workbook(excel_path, data_only=True)
    name_to_code = {}
    codekey_to_code = {}

    for sheet in wb.worksheets:
        rows = sheet.iter_rows(min_row=1)
        try:
            header = next(rows)
        except StopIteration:
            continue

        code_i = find_col_idx(header, CODE_HEADERS)
        ar_i = find_col_idx(header, AR_HEADERS)
        en_i = find_col_idx(header, EN_HEADERS)
        fn_i = find_col_idx(header, FILENAME_HEADERS)

        if not code_i:
            print(f"⚠ تخطي الورقة '{sheet.title}': لا يوجد عمود للكود", file=sys.stderr)
            continue

        for row in sheet.iter_rows(min_row=2):
            code_raw = row[code_i - 1].value
            if not code_raw:
                continue
            code = str(code_raw).strip()
            codekey = normalize_key(code)
            if codekey:
                codekey_to_code[codekey] = code  # احتفظ بالأصل كما هو

            if ar_i:
                ar_val = row[ar_i - 1].value
                if ar_val:
                    name_to_code[normalize_key(ar_val)] = code
            if en_i:
                en_val = row[en_i - 1].value
                if en_val:
                    name_to_code[normalize_key(en_val)] = code
            if fn_i:
                fn_val = row[fn_i - 1].value
                if fn_val:
                    name_to_code[normalize_key(fn_val)] = code
                    # وقد يكون الاسم مُضمّنًا بلا .pdf
                    name_to_code[normalize_key(str(fn_val).replace(".pdf", ""))] = code

    return name_to_code, codekey_to_code


def plan_renames(folder: Path, name_to_code: dict, codekey_to_code: dict):
    """
    يبني خطة إعادة التسمية: قائمة من (src, dst, السبب)
    """
    plans = []
    unmatched = []

    pdf_files = sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == PDF_EXT])

    # قائمة أكواد مطبّعة للبحث السريع
    all_code_keys = set(codekey_to_code.keys())

    for p in pdf_files:
        base = p.stem  # بدون .pdf
        base_key = normalize_key(base)

        # 1) إن كان الاسم بالفعل كودًا => تخطَّ
        if base_key in all_code_keys:
            continue

        # 2) تطابق بالاسم (عربي/إنجليزي/filename)
        code = name_to_code.get(base_key)

        # 3) محاولة أخف: لو الاسم يحوي الكود كجزء منه
        if not code:
            # مثال: "طلب_سلفة_ف-001" يحتوي الكود داخل الاسم
            for ck in all_code_keys:
                if ck and ck in base_key:
                    code = codekey_to_code[ck]
                    break

        if code:
            dst = p.with_name(f"{code}{PDF_EXT}")
            # تجنّب الكتابة فوق ملف موجود مختلف
            final_dst = dst
            n = 1
            while final_dst.exists() and final_dst.resolve() != p.resolve():
                # لو موجود مسبقًا باسم الكود، أضف لاحقة رقمية آمنة
                final_dst = p.with_name(f"{code}__{n}{PDF_EXT}")
                n += 1
            reason = "match-by-name" if name_to_code.get(base_key) else "match-by-substring"
            plans.append((p, final_dst, reason))
        else:
            unmatched.append(p)

    return plans, unmatched


def main():
    ap = argparse.ArgumentParser(description="إعادة تسمية ملفات PDF إلى الأكواد بناءً على forms.xlsx")
    ap.add_argument("--excel", default="forms.xlsx", help="مسار ملف Excel (افتراضي: forms.xlsx)")
    ap.add_argument("--folder", default=".", help="المجلد الذي يحتوي PDFs (افتراضي: المجلد الحالي)")
    ap.add_argument("--apply", action="store_true", help="تنفيذ فعلي (وإلا فسيكون Dry-Run)")
    ap.add_argument("--report", default="rename_report.csv", help="اسم تقرير CSV")
    args = ap.parse_args()

    folder = Path(args.folder).resolve()
    excel_path = Path(args.excel).resolve()

    if not folder.exists():
        print(f"❌ المجلد غير موجود: {folder}", file=sys.stderr)
        sys.exit(2)
    if not excel_path.exists():
        print(f"❌ ملف Excel غير موجود: {excel_path}", file=sys.stderr)
        sys.exit(2)

    print(f"📘 Excel: {excel_path}")
    print(f"📂 Folder: {folder}")

    name_to_code, codekey_to_code = build_mapping_from_excel(excel_path)
    print(f"🔎 خرائط: {len(name_to_code)} اسم → كود، {len(codekey_to_code)} كود معروف.")

    plans, unmatched = plan_renames(folder, name_to_code, codekey_to_code)

    # تقرير CSV
    report_path = folder / args.report
    with open(report_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["src", "dst", "reason", "status"])
        for src, dst, reason in plans:
            w.writerow([src.name, dst.name, reason, "PLANNED"])
        for p in unmatched:
            w.writerow([p.name, "", "unmatched", "SKIPPED"])

    print(f"📝 تقرير: {report_path.name}")
    print(f"✅ خطط إعادة التسمية: {len(plans)} ملف")
    print(f"⚠ غير المطابق: {len(unmatched)} ملف")

    if not args.apply:
        print("\n(Dry-Run) لم يتم أي تغيير. أعد التشغيل مع --apply للتنفيذ الفعلي.")
        return

    # تنفيذ فعلي
    applied = 0
    skipped = 0
    for src, dst, _ in plans:
        try:
            if dst.exists() and dst.resolve() == src.resolve():
                skipped += 1
                continue
            src.rename(dst)
            applied += 1
            print(f"✔ {src.name}  →  {dst.name}")
        except Exception as e:
            print(f"❌ فشل في {src.name} → {dst.name}: {e}", file=sys.stderr)

    if unmatched:
        print("\n⚠ ملفات لم يتم التعرّف عليها (حدّث الأسماء في Excel أو أعد التسمية يدويًا):")
        for p in unmatched:
            print("   -", p.name)

    print(f"\nتم. أعيدت تسمية {applied} ملفًا، وتخطّيت {skipped}.")


if __name__ == "__main__":
    main()
