# management/commands/import_forms.py
import re
from pathlib import Path
from collections import defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.apps import apps
from django.conf import settings
from django.db import transaction
from django.core.files import File

try:
    from openpyxl import load_workbook
except Exception:
    raise CommandError("openpyxl غير مثبت. أضِفه إلى requirements.txt ثم ثبّته.")

# ----------------- أدوات مساعدة -----------------
def norm(s):
    return (str(s).strip() if s is not None else "").strip()

def normalize_header(h):
    h = norm(h).lower()
    mapping = {
        "serial_number": {"serial", "serial number", "form no", "form number", "form id",
                          "code", "form code", "رقم", "الكود", "serial_number"},
        "name_ar": {"name_ar", "arabic", "arabic name", "الاسم العربي", "name ar"},
        "name_en": {"name_en", "english", "english name", "الاسم الانكليزي", "name en"},
        "category": {"category", "الفئة", "القسم الداخلي", "التصنيف"},
        "description": {"description", "وصف", "details", "ملاحظات"},
        "section": {"section", "القسم", "section_ar", "section_en", "department"},
    }
    for key, aliases in mapping.items():
        if h in aliases:
            return key
    return h

DASHES = {"\u2013", "\u2014", "_"}  # – — _
def norm_code(code: str) -> str:
    """تطبيع الكود: إزالة .pdf، توحيد الشرطات، حذف المسافات، وتحويل HR-1 إلى HR-001."""
    s = norm(code)
    if not s:
        return ""
    s = s.replace(".PDF", "").replace(".pdf", "")
    for d in DASHES:
        s = s.replace(d, "-")
    s = re.sub(r"\s+", "", s)
    s = s.lower()
    m = re.match(r"^([a-z]+)-?(\d+)$", s)
    if m:
        letters, num = m.group(1), m.group(2).zfill(3)
        return f"{letters}-{num}"
    return s

def has_fields(model, needed):
    field_names = {f.name for f in model._meta.get_fields() if hasattr(f, "name")}
    return needed.issubset(field_names)

def find_models(app_label=None):
    """إرجاع Section و FormModel من التطبيق المحدد أو بالكشف التلقائي."""
    if app_label:
        try:
            Section = apps.get_model(app_label, "Section")
            FormModel = apps.get_model(app_label, "FormModel")
            return Section, FormModel
        except LookupError as e:
            raise CommandError(f"لا يوجد app '{app_label}' يحوي Section/FormModel. ({e})")
    Section = FormModel = None
    for m in apps.get_models():
        if m.__name__ == "Section" and has_fields(m, {"name_ar", "name_en"}):
            Section = m
        if m.__name__ == "FormModel" and has_fields(
            m, {"serial_number", "name_ar", "name_en", "category", "description", "file"}
        ):
            FormModel = m
    if not Section or not FormModel:
        labels = [a.label for a in apps.get_app_configs()]
        raise CommandError(
            "تعذّر إيجاد Section/FormModel تلقائيًا. مرّر اسم التطبيق بخيار --app-label.\n"
            f"التطبيقات المتاحة: {', '.join(labels)}"
        )
    return Section, FormModel

def detect_header_row(ws, max_scan=15):
    """البحث عن صف العناوين خلال أول N صفوف باحتساب تطابق الأسماء المعروفة."""
    wanted = {"serial_number", "name_ar", "name_en", "category", "description", "section"}
    best = (0, 1, [])  # (score, row_idx, headers)
    for r in range(1, min(ws.max_row, max_scan) + 1):
        cells = [normalize_header(c.value) for c in ws[r]]
        score = sum(1 for v in cells if v in wanted)
        if score > best[0]:
            best = (score, r, cells)
    return best  # يرجع (score, header_row_idx, headers)

# ----------------- الأمر -----------------
class Command(BaseCommand):
    help = "يقرأ كل ملفات PDF في data/ ويطابقها مع صفوف forms.xlsx (كل الشيتات) ويُنشئ/يحدّث FormModel."

    def add_arguments(self, parser):
        parser.add_argument("--data-dir",
                            default=str(Path(settings.BASE_DIR) / "data"),
                            help="مسار مجلد البيانات (افتراضي BASE_DIR/data)")
        parser.add_argument("--excel",
                            default="forms.xlsx",
                            help="اسم/مسار ملف الإكسل داخل مجلد البيانات (افتراضي forms.xlsx)")
        parser.add_argument("--sheet",
                            help="إن رغبت: اسم شيت محدد داخل الإكسل بدل قراءة الكل.")
        parser.add_argument("--dry-run", action="store_true",
                            help="تشغيل تجريبي بلا حفظ")
        parser.add_argument("--app-label",
                            help="وسم تطبيق Django الذي يحوي Section وFormModel (مثل: core)")
        parser.add_argument("--create-missing-sections", action="store_true",
                            help="إنشاء الأقسام المفقودة تلقائيًا إذا لم تُوجد في قاعدة البيانات.")

    def handle(self, *args, **opts):
        data_dir = Path(opts["data_dir"]).resolve()
        excel_path = (data_dir / opts["excel"]).resolve()
        sheet_only = opts.get("sheet")
        dry_run = opts["dry_run"]
        app_label = opts.get("app_label")
        create_missing_sections = opts["create_missing_sections"]

        if not data_dir.exists():
            raise CommandError(f"مجلد البيانات غير موجود: {data_dir}")
        if not excel_path.exists():
            raise CommandError(f"ملف الإكسل غير موجود: {excel_path}")

        Section, FormModel = find_models(app_label)

        self.stdout.write(self.style.NOTICE(f"📂 DATA DIR: {data_dir}"))
        self.stdout.write(self.style.NOTICE(f"📄 EXCEL  : {excel_path.name}"))

        # 1) فهرس ملفات PDF
        pdf_index = {}
        for p in data_dir.glob("*.pdf"):
            key = norm_code(p.stem)
            if key:
                pdf_index[key] = p
        if not pdf_index:
            self.stdout.write(self.style.WARNING("لم يتم العثور على أي PDF في المجلد."))

        # 2) قراءة كل الشيتات من الإكسل مع كشف صف العناوين
        wb = load_workbook(excel_path, data_only=True)
        sheetnames = [sheet_only] if sheet_only else wb.sheetnames

        rows_data = []
        for sname in sheetnames:
            ws = wb[sname]
            score, header_row_idx, headers = detect_header_row(ws)
            if score == 0:
                # لا عناوين واضحة — تجاهل الشيت
                self.stdout.write(self.style.WARNING(f"تخطّي الشيت '{sname}' لعدم العثور على صف عناوين مناسب."))
                continue

            # قراءة البيانات أسفل صف العناوين
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if row is None:
                    continue
                row_dict = {normalize_header(headers[i]): row[i] for i in range(min(len(headers), len(row)))}
                serial = norm(row_dict.get("serial_number"))
                if not serial:
                    continue
                section_name = norm(row_dict.get("section")) or norm(sname)  # إن لم يوجد عمود قسم، استخدم اسم الشيت
                rows_data.append({
                    "serial_number": serial,
                    "serial_key": norm_code(serial),
                    "name_ar": norm(row_dict.get("name_ar")),
                    "name_en": norm(row_dict.get("name_en")),
                    "category": norm(row_dict.get("category")),
                    "description": norm(row_dict.get("description")),
                    "section": section_name,
                })

        # فهرس بحسب الكود المطَبَّع
        excel_index = {}
        for r in rows_data:
            if r["serial_key"]:
                excel_index[r["serial_key"]] = r

        self.stdout.write(self.style.NOTICE(f"🧾 Loaded {len(rows_data)} rows from {len(sheetnames)} sheet(s)."))
        self.stdout.write(self.style.NOTICE(f"📑 PDFs found: {len(pdf_index)}"))

        created = updated = skipped_no_excel = skipped_no_section = skipped_no_pdf = 0
        problems = defaultdict(list)

        @transaction.atomic
        def do_work():
            nonlocal created, updated, skipped_no_excel, skipped_no_section, skipped_no_pdf

            for key, pdf_path in pdf_index.items():
                row = excel_index.get(key)
                if not row:
                    skipped_no_excel += 1
                    problems["لا يوجد صف في الإكسل لهذا الكود"].append(pdf_path.name)
                    continue

                # ابحث/أنشئ القسم
                section_name = row.get("section")
                section_obj = (Section.objects.filter(name_ar__iexact=section_name).first()
                               or Section.objects.filter(name_en__iexact=section_name).first())
                if not section_obj:
                    if create_missing_sections and not dry_run:
                        # أنشئ قسمًا جديدًا بالاسمين نفسه مؤقتًا
                        section_obj = Section.objects.create(name_ar=section_name, name_en=section_name)
                    else:
                        skipped_no_section += 1
                        problems["القسم غير موجود في قاعدة البيانات"].append(f"{pdf_path.name} -> {section_name!r}")
                        continue

                obj = FormModel.objects.filter(serial_number__iexact=row["serial_number"]).first()

                if obj:
                    changed = False
                    if obj.section_id != section_obj.id:
                        obj.section = section_obj; changed = True
                    for fld in ("name_ar", "name_en", "category", "description"):
                        new_val = row.get(fld, "")
                        if getattr(obj, fld) != new_val:
                            setattr(obj, fld, new_val); changed = True
                    filename = pdf_path.name
                    if not obj.file or Path(obj.file.name).name != filename:
                        if not dry_run:
                            with open(pdf_path, "rb") as fh:
                                obj.file.save(filename, File(fh), save=False)
                        changed = True
                    if changed and not dry_run:
                        obj.save()
                    if changed:
                        updated += 1
                else:
                    obj = FormModel(
                        section=section_obj,
                        serial_number=row["serial_number"],
                        name_ar=row.get("name_ar", ""),
                        name_en=row.get("name_en", ""),
                        category=row.get("category", ""),
                        description=row.get("description", ""),
                    )
                    if not dry_run:
                        with open(pdf_path, "rb") as fh:
                            obj.file.save(pdf_path.name, File(fh), save=False)
                        obj.save()
                    created += 1

            # صفوف لا PDF لها
            for key, r in excel_index.items():
                if key not in pdf_index:
                    skipped_no_pdf += 1
                    problems["لا يوجد PDF لهذا الكود"].append(r["serial_number"])

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — لن يتم أي حفظ."))
        do_work()

        # ملخص
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(f"✅ Created: {created}"))
        self.stdout.write(self.style.SUCCESS(f"✅ Updated: {updated}"))
        self.stdout.write(self.style.WARNING(f"⛔ Skipped (no excel row): {skipped_no_excel}"))
        self.stdout.write(self.style.WARNING(f"⛔ Skipped (section missing): {skipped_no_section}"))
        self.stdout.write(self.style.WARNING(f"⛔ Skipped (no PDF for excel row): {skipped_no_pdf}"))

        if problems:
            self.stdout.write("\nتفاصيل المشاكل:")
            for k, items in problems.items():
                for it in items[:50]:
                    self.stdout.write(f" - {k}: {it}")
            if any(len(v) > 50 for v in problems.values()):
                self.stdout.write("... (تم تقصير القائمة)")
