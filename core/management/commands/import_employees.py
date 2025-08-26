from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from core.models import Section, UserSectionPermission
import openpyxl

User = get_user_model()

class Command(BaseCommand):
    help = "Import sections and users from Excel files"

    def handle(self, *args, **kwargs):
        self.import_sections()
        self.import_users()

    def import_sections(self):
        self.stdout.write("📁 Importing sections from sections.xlsx...")
        try:
            wb = openpyxl.load_workbook("sections.xlsx")
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR("❌ File 'sections.xlsx' not found."))
            return

        ws = wb.active
        created, skipped = 0, 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name_ar, name_en = row

            if not name_ar or not name_en:
                self.stdout.write(self.style.WARNING("⚠️ Skipped section with missing data."))
                skipped += 1
                continue

            if Section.objects.filter(name_ar=name_ar, name_en=name_en).exists():
                skipped += 1
                continue

            Section.objects.create(name_ar=name_ar, name_en=name_en)
            self.stdout.write(self.style.SUCCESS(f"✅ Created section: {name_en}"))
            created += 1

        self.stdout.write(self.style.SUCCESS(f"✔️ Sections import completed: {created} created, {skipped} skipped."))

    def import_users(self):
        self.stdout.write("\n📁 Importing users from employees.xlsx...")
        try:
            wb = openpyxl.load_workbook("employees.xlsx")
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR("❌ File 'employees.xlsx' not found."))
            return

        ws = wb.active
        created, skipped = 0, 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            username, password, role = row

            if not username or not password or not role:
                self.stdout.write(self.style.WARNING(f"⚠️ Skipped: Incomplete data for '{username}'"))
                skipped += 1
                continue

            if User.objects.filter(username=username).exists():
                self.stdout.write(self.style.WARNING(f"⚠️ Skipped: Username '{username}' already exists."))
                skipped += 1
                continue

            is_staff = True if role.lower() in ['manager', 'hr'] else False
            user = User.objects.create_user(username=username, password=password, is_staff=is_staff,  role=role.lower() )
            self.stdout.write(self.style.SUCCESS(f"✅ Created: {username} — {role}"))
            created += 1

            # ✅ منح صلاحية الاطلاع على جميع الأقسام إذا كان المدير أو HR
            if role.lower() in ['manager', 'hr']:
                all_sections = Section.objects.all()
                for section in all_sections:
                    UserSectionPermission.objects.create(user=user, section=section)
                self.stdout.write(self.style.SUCCESS(f"🟢 Permissions granted for all sections to {username}"))

        self.stdout.write(self.style.SUCCESS(f"✔️ Users import completed: {created} created, {skipped} skipped."))
